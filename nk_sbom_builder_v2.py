import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from  parser.package_lock_parser_v2  import parse_package_lock 

def get_parent_directory():
    # 获取当前工作目录
    current_dir = os.getcwd()
    # 获取当前目录的父目录
    return os.path.dirname(current_dir)


def get_current_directory():
    # 获取当前工作目录
    return os.getcwd()

def prepare_data_for_excel(data_list, headers):
    # 准备数据写入 Excel
    excel_data = [headers]
    for data in data_list:
        excel_data.append([data[field] for field in headers])
    return excel_data

def write_data_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    for item in data:
        ws.append(item)

    # 设置表头字体样式
    for cell in ws[1]:
        cell.font = Font(size=14, bold=True)

    # 设置 name 列的单元格宽度为 200
    ws.column_dimensions[get_column_letter(1)].width = 4
    ws.column_dimensions[get_column_letter(2)].width = 45
    ws.column_dimensions[get_column_letter(3)].width = 30
    ws.column_dimensions[get_column_letter(4)].width = 8
    ws.column_dimensions[get_column_letter(5)].width = 96
    ws.column_dimensions[get_column_letter(6)].width = 40
    ws.column_dimensions[get_column_letter(7)].width = 17
    ws.column_dimensions[get_column_letter(8)].width = 22
    ws.column_dimensions[get_column_letter(9)].width = 26
    ws.column_dimensions[get_column_letter(10)].width = 5
    ws.column_dimensions[get_column_letter(11)].width = 6

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 添加时间戳到文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name, extension = os.path.splitext(output_file)
    excel_file = f"{file_name}_{timestamp}{extension}"

    wb.save(excel_file)
    print(f"Excel 文件已保存至：{excel_file}")

def process_build(MAX_COUNT=0):
    nk_package_data=[]

    # 构造 package.lock 文件路径
    parent_dir = get_current_directory()
    package_lock_file = os.path.join(parent_dir, "files", "package-lock.json")

    # 解析 package.lock 文件并打印包信息
    _, package_list, _ = parse_package_lock(package_lock_file,MAX_COUNT)

    id=1
    misCount=0
    dep_map={}
    for package in package_list: 
        name = package.get('name','')  
        author=package.get('author','') 

        nk_package_info = {
            'ID': id,
            'Supplier Name': author,
            'Component Name': name,
            'Version of the Component': package.get('version', '-'),
            'Other Unique Identifiers':package.get('integrity', '-'),
            'Hash of the Component (Optional)': package.get('shasum', ''), 
            'peerDependencies': package.get('peerDependencies', []),
            'Author of SBOM Data': author,
            'Timestamp': package.get('publishTime',''),  
            'Software Category':'OTS', 
            'License Declared':package.get('license', '-'), 
            'Comment':''
        } 

        nk_package_data.append(nk_package_info)
        if name not in dep_map:
            dep_map[name]=f'include in id#{id}'
        id+=1

    # 更新关系
    for nk_pack_info in nk_package_data:
        new_deps=[]
        for dep in nk_pack_info.get('peerDependencies',[]):
            new_deps.append(dep_map.get(dep,dep))
        nk_pack_info['Dependency Relationship']=','.join(new_deps)

    # 准备数据写入 Excel
    head_data = ["ID","Supplier Name", "Component Name", "Version of the Component", "Other Unique Identifiers",
    "Hash of the Component (Optional)","Dependency Relationship","Author of SBOM Data","Timestamp",
    "Software Category","License Declared","Comment"]
    nk_package_excel_data = prepare_data_for_excel(nk_package_data, head_data)

    # 写入 Excel
    excel_file = os.path.join(parent_dir, "_output", "MDIS-APP_软件物料清单.xlsx")
    write_data_to_excel(nk_package_excel_data, excel_file)  
    print(f"APP_软件物料清单 已经保存至：{excel_file}")

def main(): 
    process_build(0) 

if __name__ == "__main__":
    main()
