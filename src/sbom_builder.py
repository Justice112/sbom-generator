import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from src.parser.package_lock_parser_v2 import parse_package_lock

def prepare_data_for_excel(data_list):
    """
    准备数据写入 Excel。

    Args:
        data_list (list): 数据列表，每个元素是一个字典，表示一个数据记录。

    Returns:
        list: 准备好的数据列表，用于写入 Excel。
    """
    if not data_list:
        raise ValueError("数据列表为空，无法准备数据写入 Excel。")

    headers = list(data_list[0].keys())

    # 检查所有字典是否具有相同的键
    for data in data_list:
        if not isinstance(data, dict):
            raise TypeError("数据列表中的元素不是字典。")
        if set(data.keys()) != set(headers):
            raise ValueError("所有字典必须具有相同的键。")

    excel_data = [headers]
    for data in data_list:
        row = [data[field] for field in headers]
        excel_data.append(row)

    return excel_data


def write_data_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    for item in data:
        ws.append(item)
    
    set_header_style(ws)
    set_column_width(ws)

    ws.freeze_panes = 'A2'

    excel_file = add_timestamp_to_filename(output_file)
    wb.save(excel_file)
    print(f"Excel 文件已保存至：{excel_file}")
    
    return excel_file

def set_header_style(ws):
    for cell in ws[1]:
        cell.font = Font(size=14, bold=True)

def set_column_width(ws):
    column_widths = {
        1: 4, 2: 45, 3: 30, 4: 8, 5: 96, 6: 40, 7: 17, 8: 22, 9: 26, 10: 5, 11: 6
    }
    for column, width in column_widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width

def add_timestamp_to_filename(output_file):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name, extension = os.path.splitext(output_file)
    return f"{file_name}_{timestamp}{extension}"

def build_package_info(package_list):
    package_data = []
    id = 1
    dep_map = {}

    for package in package_list:
        name = package.get('name', '')
        author = package.get('author', '')

        package_info = {
            'ID': id,
            'Supplier Name': author,
            'Component Name': name,
            'Version of the Component': package.get('version', '-'),
            'Other Unique Identifiers': package.get('integrity', '-'),
            'Hash of the Component (Optional)': package.get('shasum', '-'),
            'peerDependencies': package.get('peerDependencies', []),
            'Dependency Relationship':'',
            'Author of SBOM Data': author,
            'Timestamp': package.get('publishTime', '-'),
            'Software Category': 'OTS',
            'License Declared': package.get('license', '-'),
            'Comment': ''
        }

        package_data.append(package_info)
        if name not in dep_map:
            dep_map[name] = f'include in id#{id}'
        id += 1

    for pack_info in package_data:
        new_deps = []
        for dep in pack_info.get('peerDependencies', []):
            new_deps.append(dep_map.get(dep, dep))
        pack_info['Dependency Relationship'] = ','.join(new_deps) 
    
    # TODO 删除 中 peerDependencies 字段
    for pack_info in package_data:
        del pack_info['peerDependencies']  # 删除 peerDependencies 字段

    return package_data

def process_build(package_lock_path, output_file, max_count=0):
    """
    处理构建流程，生成软件物料清单 Excel 文件。

    Args:
        package_lock_path (str): package-lock.json 文件的路径。
        output_file (str): Excel 文件保存路径及名称。
        max_count (int, optional): 解析的最大包数。默认为 0，表示解析全部。
    """
    excel_file=''
    
    if not os.path.exists(package_lock_path):
        print(f"错误：文件 {package_lock_path} 不存在")
        return excel_file
    
    if not os.path.exists(os.path.dirname(output_file)):
        print(f"错误：输出目录 {os.path.dirname(output_file)} 不存在")
        return excel_file
    
    try:
        _, package_list, _ = parse_package_lock(package_lock_path, max_count)
        package_data = build_package_info(package_list)
        excel_data = prepare_data_for_excel(package_data)
        excel_file=write_data_to_excel(excel_data, output_file)
        print(f"APP_软件物料清单 已经保存至：{output_file}")
    except FileNotFoundError:
        print(f"错误：找不到文件 {package_lock_path}")
    except Exception as e:
        print(f"错误：处理构建流程时发生异常：{e}")
    
    return excel_file
