import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from parser.package_lock_parser_v2 import parse_package_lock

def get_parent_directory():
    return os.path.dirname(os.getcwd())

def prepare_data_for_excel(data_list):
    """
    准备数据写入 Excel。

    Args:
        data_list (list): 数据列表，每个元素是一个字典，表示一个数据记录。

    Returns:
        list: 准备好的数据列表，用于写入 Excel。
    """
    # 提取第一个数据记录的字段作为表头
    headers = list(data_list[0].keys())

    # 将数据列表转换为二维列表，用于写入 Excel
    excel_data = [headers]
    for data in data_list:
        excel_data.append([data[field] for field in headers])

    return excel_data

def write_data_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active
    for item in data:
        ws.append(item)
    
    set_header_style(ws)
    set_column_width(ws)

    ws.freeze_panes = 'A2'

    excel_file = add_timestamp_to_filename(output_file)
    wb.save(excel_file)
    print(f"Excel 文件已保存至：{excel_file}")

def set_header_style(ws):
    for cell in ws[1]:
        cell.font = Font(size=14, bold=True)

def set_column_width(ws):
    column_widths = {
        1: 4, 2: 45, 3: 30, 4: 8, 5: 96, 6: 40, 7: 17, 8: 22, 9: 26, 10: 5, 11: 6
    }
    for column, width in column_widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width

def add_timestamp_to_filename(output_file):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name, extension = os.path.splitext(output_file)
    return f"{file_name}_{timestamp}{extension}"

def process_build(max_count=0):
    """
    处理构建流程，生成软件物料清单 Excel 文件。

    Args:
        max_count (int, optional): 解析的最大包数。默认为 0，表示解析全部。
    """
    
    package_data = []
    parent_dir = get_parent_directory()
    package_lock_file = os.path.join(parent_dir, "files", "package-lock.json")
    _, package_list, _ = parse_package_lock(package_lock_file, max_count)

    id = 1
    dep_map = {}
    for package in package_list:
        name = package.get('name', '')
        author = package.get('author', '')

        package_info = {
            'ID': id,
            'Supplier Name': author,
            'Component Name': name,
            'Version of the Component': package.get('version', '-'),
            'Other Unique Identifiers': package.get('integrity', '-'),
            'Hash of the Component (Optional)': package.get('shasum', ''),
            'peerDependencies': package.get('peerDependencies', []),
            'Author of SBOM Data': author,
            'Timestamp': package.get('publishTime', ''),
            'Software Category': 'OTS',
            'License Declared': package.get('license', '-'),
            'Comment': ''
        }

        package_data.append(package_info)
        if name not in dep_map:
            dep_map[name] = f'include in id#{id}'
        id += 1

    for pack_info in package_data:
        new_deps = []
        for dep in pack_info.get('peerDependencies', []):
            new_deps.append(dep_map.get(dep, dep))
        pack_info['Dependency Relationship'] = ','.join(new_deps)

    package_excel_data = prepare_data_for_excel(package_data)

    excel_file = os.path.join(parent_dir, "_output", "MDIS-APP_软件物料清单.xlsx")
    write_data_to_excel(package_excel_data, excel_file)
    print(f"APP_软件物料清单 已经保存至：{excel_file}")

def main():
    process_build()

if __name__ == "__main__":
    main()
